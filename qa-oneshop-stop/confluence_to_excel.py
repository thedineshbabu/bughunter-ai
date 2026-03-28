"""Fetch a Confluence page and export its tables to an Excel workbook.

Usage:
    python confluence_to_excel.py <page_url_or_id> [output_filename.xlsx]

Example:
    python confluence_to_excel.py https://kornferry.atlassian.net/wiki/spaces/KO/pages/2464874582/KF+One+RBAC+Definition
"""

from __future__ import annotations

import re
import sys
from pathlib import Path

import requests
from bs4 import BeautifulSoup, Tag
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from requests.auth import HTTPBasicAuth

from config import Config


def extract_page_id(url_or_id: str) -> str:
    if url_or_id.isdigit():
        return url_or_id
    m = re.search(r"/pages/(\d+)", url_or_id)
    if m:
        return m.group(1)
    m = re.search(r"pageId=(\d+)", url_or_id)
    if m:
        return m.group(1)
    raise ValueError(f"Cannot extract page ID from: {url_or_id}")


def fetch_page_html(page_id: str) -> tuple[str, str]:
    """Return (title, body_html) for a Confluence page."""
    base = Config.CONFLUENCE_BASE_URL.rstrip("/")
    auth = HTTPBasicAuth(Config.CONFLUENCE_EMAIL, Config.CONFLUENCE_API_TOKEN)
    url = f"{base}/rest/api/content/{page_id}"
    resp = requests.get(
        url,
        auth=auth,
        headers={"Accept": "application/json"},
        params={"expand": "body.storage"},
        timeout=30,
    )
    resp.raise_for_status()
    data = resp.json()
    title = data.get("title", "Untitled")
    body_html = data.get("body", {}).get("storage", {}).get("value", "")
    return title, body_html


def parse_tables(html: str) -> list[dict]:
    """Parse all <table> elements from HTML into a list of dicts.

    Each dict has:
      - heading: the nearest <h*> text before the table (or "Table N")
      - rows: list[list[str]] — cell text per row
    """
    soup = BeautifulSoup(html, "html.parser")
    tables: list[dict] = []

    for idx, table_tag in enumerate(soup.find_all("table"), start=1):
        heading = _find_preceding_heading(table_tag, idx)
        rows: list[list[str]] = []

        for tr in table_tag.find_all("tr"):
            cells = tr.find_all(["th", "td"])
            row = [_cell_text(c) for c in cells]
            if any(cell.strip() for cell in row):
                rows.append(row)

        if rows:
            tables.append({"heading": heading, "rows": rows})

    return tables


def _find_preceding_heading(tag: Tag, fallback_idx: int) -> str:
    """Walk previous siblings / parent to find the nearest heading."""
    node = tag
    for _ in range(20):
        prev = node.find_previous_sibling(re.compile(r"^h[1-6]$"))
        if prev:
            return prev.get_text(strip=True)
        node = node.parent
        if node is None:
            break
    return f"Table {fallback_idx}"


def _cell_text(cell: Tag) -> str:
    """Extract clean text from a table cell, collapsing whitespace."""
    for br in cell.find_all("br"):
        br.replace_with("\n")
    text = cell.get_text(separator=" ").strip()
    return re.sub(r"[ \t]+", " ", text)


def write_excel(tables: list[dict], title: str, output_path: Path) -> None:
    """Write parsed tables into a styled Excel workbook."""
    wb = Workbook()
    wb.remove(wb.active)

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    cell_font = Font(size=10)
    cell_align = Alignment(vertical="top", wrap_text=True)

    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    alt_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    for tbl in tables:
        sheet_name = _safe_sheet_name(tbl["heading"])
        ws = wb.create_sheet(title=sheet_name)

        for r_idx, row in enumerate(tbl["rows"], start=1):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.border = thin_border

                if r_idx == 1:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_align
                else:
                    cell.font = cell_font
                    cell.alignment = cell_align
                    if r_idx % 2 == 0:
                        cell.fill = alt_fill

        # Auto-size columns
        for col_idx in range(1, ws.max_column + 1):
            max_len = 10
            for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, values_only=True):
                val = str(row[0] or "")
                longest_line = max((len(line) for line in val.split("\n")), default=0)
                if longest_line > max_len:
                    max_len = longest_line
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 60)

        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = "A2"

    if not wb.sheetnames:
        ws = wb.create_sheet(title="No Tables Found")
        ws.cell(row=1, column=1, value="No tables were found on this Confluence page.")

    wb.save(str(output_path))


def _safe_sheet_name(name: str) -> str:
    """Sanitise a string for use as an Excel sheet name (max 31 chars)."""
    name = re.sub(r'[\\/*?\[\]:]', " ", name).strip()
    return name[:31] if name else "Sheet"


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)

    page_input = sys.argv[1]
    page_id = extract_page_id(page_input)

    print(f"Fetching Confluence page {page_id} ...")
    title, html = fetch_page_html(page_id)
    print(f"Page title: {title}")

    tables = parse_tables(html)
    print(f"Found {len(tables)} table(s)")

    for t in tables:
        print(f"  - {t['heading']}: {len(t['rows'])} rows x {max(len(r) for r in t['rows'])} cols")

    output_dir = Path(Config.OUTPUT_DIR)
    output_dir.mkdir(parents=True, exist_ok=True)
    safe_title = re.sub(r'[^\w\s-]', '', title).strip().replace(' ', '_')[:60]
    output_name = sys.argv[2] if len(sys.argv) > 2 else f"{safe_title}.xlsx"
    output_path = output_dir / output_name

    write_excel(tables, title, output_path)
    print(f"\nExcel saved to: {output_path}")


if __name__ == "__main__":
    main()
